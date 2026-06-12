"""Normalize arbitrary chapter XLSX/CSV tables into the recommended input schema.

Real-world chapter exports rarely use the exact column names that
:func:`tcm_fuzzywiki.io.read_chapters` recommends.  This module maps common
Chinese/English column aliases onto the V5.0 schema, guesses the main text
column when no alias matches, fills documented defaults, and writes a JSON
column-mapping report so the normalization itself stays auditable.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils.dataframe import dataframe_to_rows

COLUMN_ALIASES: dict[str, list[str]] = {
    "source_id": ["source_id", "id", "编号", "章节id", "条目id"],
    "book_name": ["book_name", "书名", "古籍名称", "书籍", "book", "title"],
    "volume_name": ["volume_name", "卷名", "卷", "volume", "册"],
    "chapter_title": ["chapter_title", "章节标题", "章节", "篇名", "章名", "chapter", "section_title", "标题"],
    "chapter_order": ["chapter_order", "章节序号", "序号", "order", "chapter_index", "index"],
    "dynasty": ["dynasty", "朝代", "年代"],
    "author": ["author", "作者"],
    "text_original": ["text_original", "original_text", "原文", "正文", "内容", "chapter_text", "text", "古籍原文", "分章内容", "段落"],
    "text_punctuated": ["text_punctuated", "标点原文", "点校文本"],
    "text_modern": ["text_modern", "现代文", "白话文", "译文"],
    "chapter_summary": ["chapter_summary", "摘要", "小结", "summary"],
    "text_type": ["text_type", "文本类型", "类型"],
    "topic_hint": ["topic_hint", "主题", "主题提示", "topic"],
    "school_tag": ["school_tag", "流派", "学派"],
    "region_tag": ["region_tag", "地域", "地区"],
    "tradition_id": ["tradition_id", "学术传统", "传统"],
    "text_family": ["text_family", "文献类型家族", "文献类型"],
    "citation_family": ["citation_family", "引用谱系", "版本谱系"],
    "notes": ["notes", "备注", "说明"],
}

TEXT_DEFAULTS: dict[str, str] = {
    "volume_name": "uncertain",
    "dynasty": "uncertain",
    "author": "uncertain",
    "text_type": "古籍章节",
    "topic_hint": "中医古籍",
    "notes": "",
    "school_tag": "uncertain",
    "region_tag": "uncertain",
    "tradition_id": "uncertain",
    "text_family": "古籍章节",
    "citation_family": "uncertain",
    "text_punctuated": "",
    "text_modern": "",
    "chapter_summary": "",
}

QUALITY_DEFAULTS: dict[str, float] = {"source_authority": 0.8, "text_integrity": 0.8, "semantic_clarity": 0.8}

ORDERED_COLUMNS = [
    "source_id", "book_name", "volume_name", "chapter_title", "chapter_order",
    "dynasty", "author", "text_original", "text_punctuated", "text_modern",
    "chapter_summary", "text_type", "topic_hint", "notes",
    "school_tag", "region_tag", "tradition_id", "text_family", "citation_family",
    "source_authority", "text_integrity", "semantic_clarity",
]


def _clean_excel_cell(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, (list, dict, tuple, set)):
        value = json.dumps(value, ensure_ascii=False)
    if hasattr(value, "item"):
        try:
            value = value.item()
        except (AttributeError, ValueError):
            pass
    if isinstance(value, str):
        value = ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def safe_write_excel(frame: pd.DataFrame, output_path: str | Path, sheet_name: str = "Sheet1") -> Path:
    """Write XLSX via openpyxl directly, avoiding pandas/openpyxl writer incompatibilities."""

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = str(sheet_name)[:31] or "Sheet1"
    for row in dataframe_to_rows(frame, index=False, header=True):
        sheet.append([_clean_excel_cell(value) for value in row])
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    workbook.save(output_path)
    return output_path


def _pick_column(frame: pd.DataFrame, candidates: list[str]) -> Any:
    normalized = {str(column).strip().lower(): column for column in frame.columns}
    for candidate in candidates:
        key = candidate.strip().lower()
        if key in normalized:
            return normalized[key]
    for candidate in candidates:
        key = candidate.strip().lower()
        for column in frame.columns:
            if key and key in str(column).strip().lower():
                return column
    return None


def _guess_long_text_column(frame: pd.DataFrame) -> Any:
    columns = [column for column in frame.columns if frame[column].dtype == "object"] or list(frame.columns)
    scores = []
    for column in columns:
        series = frame[column].dropna().astype(str)
        if len(series):
            scores.append((series.str.len().mean(), series.str.len().max(), str(column), column))
    return max(scores)[3] if scores else None


def normalize_chapter_table(
    input_path: str | Path,
    output_path: str | Path,
    mapping_report_path: str | Path | None = None,
    sheet_name: int | str = 0,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Normalize an arbitrary chapter table; returns the frame and the mapping report.

    Writes the normalized XLSX, a sibling ``.csv``, and a JSON column-mapping
    report (defaulting to ``<output>.column_mapping.json``).
    """

    input_path = Path(input_path)
    output_path = Path(output_path)
    report_path = Path(mapping_report_path) if mapping_report_path else output_path.with_suffix(".column_mapping.json")

    if input_path.suffix.lower() in {".xlsx", ".xls"}:
        frame = pd.read_excel(input_path, sheet_name=sheet_name)
    elif input_path.suffix.lower() == ".csv":
        frame = pd.read_csv(input_path)
    else:
        raise ValueError(f"Unsupported input format: {input_path.suffix}")

    report: dict[str, Any] = {
        "input_path": str(input_path),
        "output_path": str(output_path),
        "sheet_name": str(sheet_name),
        "original_columns": [str(column) for column in frame.columns],
        "column_mapping": {},
    }

    out = pd.DataFrame()
    for target, aliases in COLUMN_ALIASES.items():
        column = _pick_column(frame, aliases)
        if column is not None:
            out[target] = frame[column]
            report["column_mapping"][target] = str(column)

    if "text_original" not in out.columns:
        guessed = _guess_long_text_column(frame)
        if guessed is None:
            raise ValueError(
                "无法识别正文列。请将正文列命名为 text_original/原文/正文/内容/chapter_text/text/古籍原文/分章内容/段落 之一。"
            )
        out["text_original"] = frame[guessed]
        report["column_mapping"]["text_original"] = f"{guessed}  # guessed_longest_text_column"

    n = len(frame)
    if "source_id" not in out.columns:
        out["source_id"] = [f"SRC_{i + 1:06d}" for i in range(n)]
    else:
        out["source_id"] = out["source_id"].fillna("").astype(str)
        blank = out["source_id"].str.strip().eq("")
        if blank.any():
            out.loc[blank, "source_id"] = [f"SRC_{i + 1:06d}" for i in np.flatnonzero(blank.to_numpy())]

    if "book_name" not in out.columns:
        out["book_name"] = input_path.stem
    out["book_name"] = out["book_name"].fillna(input_path.stem).astype(str).str.strip().replace("", input_path.stem)

    if "chapter_title" not in out.columns:
        out["chapter_title"] = ""
    out["chapter_title"] = out["chapter_title"].fillna("").astype(str).str.strip()
    blank_title = out["chapter_title"].eq("")
    if blank_title.any():
        out.loc[blank_title, "chapter_title"] = [f"chapter_{i + 1:06d}" for i in np.flatnonzero(blank_title.to_numpy())]

    if "chapter_order" not in out.columns:
        out["chapter_order"] = range(1, n + 1)
    out["chapter_order"] = pd.to_numeric(out["chapter_order"], errors="coerce")
    missing_order = out["chapter_order"].isna()
    if missing_order.any():
        out.loc[missing_order, "chapter_order"] = range(1, int(missing_order.sum()) + 1)
    out["chapter_order"] = out["chapter_order"].astype(int)

    for column, default in TEXT_DEFAULTS.items():
        if column not in out.columns:
            out[column] = default
        else:
            out[column] = out[column].fillna(default)
    for column, default in QUALITY_DEFAULTS.items():
        if column not in out.columns:
            out[column] = default

    out["text_original"] = out["text_original"].fillna("").astype(str).str.strip()
    before = len(out)
    out = out[out["text_original"].str.len() > 0].copy()
    report["dropped_empty_text_rows"] = int(before - len(out))
    report["final_rows"] = int(len(out))

    out = out[[c for c in ORDERED_COLUMNS if c in out.columns] + [c for c in out.columns if c not in ORDERED_COLUMNS]]

    duplicate_ids = out["source_id"][out["source_id"].duplicated()].unique().tolist()
    report["duplicate_source_ids"] = [str(value) for value in duplicate_ids]

    safe_write_excel(out, output_path)
    csv_path = output_path.with_suffix(".csv")
    out.to_csv(csv_path, index=False, encoding="utf-8-sig")
    report["output_csv"] = str(csv_path)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    report["mapping_report_path"] = str(report_path)
    return out, report
